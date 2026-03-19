"""
엑셀 종목 목록을 Supabase stock_list 테이블에 업로드하는 스크립트
사용: python scripts/upload_stock_list.py
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from app.database import get_supabase

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "..", "stock_list_20260319.xlsx")


def parse_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[wb.sheetnames[0]]

    stocks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[1]).strip() if row[1] else ""
        name = str(row[3]).strip() if row[3] else ""  # 한글 종목약명
        full_name = str(row[2]).strip() if row[2] else ""  # 한글 종목명
        market_raw = str(row[6]).strip() if row[6] else ""

        if not code or not name:
            continue

        if "KOSPI" in market_raw:
            market = "KOSPI"
        elif "KOSDAQ" in market_raw:
            market = "KOSDAQ"
        else:
            continue

        stocks.append({
            "code": code,
            "name": name,
            "full_name": full_name,
            "market": market,
        })

    return stocks


def upload(stocks):
    db = get_supabase()

    # 기존 데이터 삭제
    db.table("stock_list").delete().neq("code", "").execute()
    print(f"기존 데이터 삭제 완료")

    # 100개씩 배치 삽입
    batch_size = 100
    for i in range(0, len(stocks), batch_size):
        batch = stocks[i:i + batch_size]
        db.table("stock_list").insert(batch).execute()
        print(f"  {i + len(batch)} / {len(stocks)} 업로드됨")

    print(f"완료: {len(stocks)}개 종목 업로드")


if __name__ == "__main__":
    stocks = parse_excel()
    print(f"엑셀에서 {len(stocks)}개 종목 파싱 완료")
    upload(stocks)
